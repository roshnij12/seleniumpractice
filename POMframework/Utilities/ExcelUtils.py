import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


def get_data_from_sheet(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    testdatasheet_name = workbook[sheet_name]
    max_row = testdatasheet_name.max_row
    max_col = testdatasheet_name.max_column
    final_list = []

    for row in range(2, max_row+1):
        row_list = []
        for col in range(1, max_col+1):
            row_list.append(testdatasheet_name.cell(row, col).value)
        final_list.append(row_list)